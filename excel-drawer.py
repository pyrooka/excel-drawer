#!/usr/bin/env python

# BN 2017

import argparse
import os
import threading

import openpyxl
from openpyxl.styles import PatternFill
from PIL import Image


def get_args():
    """
    Get the given parameters from the command line.
    :return: A dictionary with the given arguments.
    """
    # Create the parser object.
    parser = argparse.ArgumentParser()

    # Add necessary arguments.
    parser.add_argument('image', help='the image you want to transform into the xlsx')
    parser.add_argument('output', help='output file')

    # Add optional arguments.
    parser.add_argument('--threads-count', type=int, help='set the number of the working threads')

    # Get the arguments.
    args = parser.parse_args()

    return args


def save_workbook(workbook, path):
    """
    Save the given workbook to the path.
    :param workbook: workbook object you want to save
    :param path: string path to the save destionation
    :return: bool: true if saved successfully, else exit
    """

    try:
        workbook.save(filename=path)
        return True
    except Exception as err:
        print('Cannot save the XLSX to the given destionation:', err, 'Please try again!')
        exit(1)


def create_workbook(image, path, threads_count=1):
    """
    Create the workbook then save it.
    :param threads_count: optional threads count for multithreading (int)
    :param image: PIL image object.
    :param path: Destination for saving.
    :return: bool: was successful (True) or not (False)
    """

    # Create a new workbook.
    wb = openpyxl.Workbook()

    # Get the active sheet.
    sheet = wb.active

    # Image size for iteration.
    image_size = image.size

    # Access to the pixels of the image.
    pixels = image.load()

    # Modify the columns width to get square like cells.
    sheet.sheet_format.baseColWidth = 2

    # Let's divide the cols based on the threads count.
    cols_per_thread = (image.size[0] // threads_count) + 1

    # Okay. Now we create an inner function what we call from all threads.
    def draw(col_from, col_to):
        # Iterates over the image pixels.
        for col in range(col_from, col_to):
            print(threading.current_thread().getName(), 'COL:', col)
            for row in range(1, image_size[1]):
                # Get only the rgb values. We don't need alpha if have.
                pixel = (pixels[col - 1, row - 1][0], pixels[col - 1, row - 1][1], pixels[col - 1, row - 1][2])
                # Get the hex color code from the RGB.
                hex_color = '%02x%02x%02x' % pixel
                sheet.cell(column=col, row=row).fill = PatternFill(start_color=hex_color, end_color=hex_color,
                                                                   fill_type='solid')

    # Store the threads.
    threads = []

    # Now create them.
    for i in range(0, threads_count):
        # The number of the first and last column for the thread. If the last is higher than the image
        # use the width of the image instead.
        col_from = 1 + i * cols_per_thread
        col_to = col_from + cols_per_thread
        col_to = col_to if col_to < image.size[0] else image.size[0] - col_from

        thread = threading.Thread(target=draw, args=(col_from, col_to))
        thread.setName('Thread#' + str(i + 1))
        threads.append(thread)

        print(thread.getName(), 'has created.')

    # After the creation start them.
    for thread in threads:
        thread.start()
        print(thread.getName(), 'has started.')

    # Let wait for the threads finish the running before continue.
    for thread in threads:
        thread.join()

    print('Saving the created document...')
    save_workbook(wb, path)


def get_image(path):
    """
    Read the image from the given path if exists.
    :param path: string path to the image (e.g. /foo/bar/image.jpg)
    :return: PIL image object
    """

    # Check if the picture exists or not.
    if not os.path.isfile(path):
        print('Cannot open the image. Please try again!')
        exit(1)

    try:
        # Open the image.
        image = Image.open(path)

        # If everything is okay return it.
        return image
    # If an error occurred.
    except Exception as err:
        print('Error occurred while trying to open the image:', err, 'Please try again!')
        exit(1)


def main():
    """
    Simply the main.
    :return: 
    """

    # Get the arguments.
    args = get_args()

    # Get the image.
    image = get_image(args.image)

    # Check the input threads count.
    if args.threads_count is None:
        pass

    threads_count = 1 if args.threads_count is None else args.threads_count

    # Initialize a workbook and sheet for coloring.
    create_workbook(image, args.output, threads_count=threads_count)

    print('Done.')


if __name__ == '__main__':
    main()
